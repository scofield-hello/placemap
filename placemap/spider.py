# -*- coding:utf-8 -*-
import os
import re
import sys
import time
import queue
import platform
import threading
from random import randint
from typing import List, Dict, NewType, Tuple, Optional
from selenium import webdriver
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.remote.webelement import WebElement
from openpyxl import Workbook, load_workbook

start_url = "https://placesmap.net/"
Country = NewType("Country", Tuple[str, str])
Area = NewType("Area", Tuple[str, str])
Position = NewType("Position", Tuple[str, str, str, str, str, str, str, str])
xls_file_path = "./车站.xlsx"
page_pattern = "<b>Found: ([0-9,]{1,10}) Places, ([0-9,]{1,10}) Pages</b>"
p_xpath = '//div[@class="six columns"]/p'
cordinate_pattern = r"([0-9\.-]{3,}),\s+([0-9\.-]{3,})"
phone_pattern = r"Phone:\s?((\(\d{1,4}\)\s?[0-9\s-]{3,30})|([+0-9\s-]{3,30}))"
web_pattern = r"\(([a-zA-Z]{1,}.*)\)"
process__queue = queue.Queue()
is_finishing = False


def start_crawl(**settings) -> None:
    countrys: List[str] = settings["COUNTRY"]
    matches: List[str] = settings["MATCHE"]
    columns: List[str] = settings["COLUMN"]
    process_task = __start_process_task(countrys, columns)
    crawl_task = __start_crawl_task(countrys, matches)
    crawl_task.start()
    process_task.start()
    process_task.join()
    print("爬虫执行结束")


def __start_crawl_task(countrys: List[str],
                       matches: List[str]) -> threading.Thread:
    thread = threading.Thread(target=__init_crawl_job,
                              name="cralw_task_0",
                              args=(countrys, matches))
    return thread


def __init_chrome_dirver() -> WebDriver:
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    driver_path = "chromedriver.exe"
    if platform.system() == "Darwin":
        driver_path = "/Users/Nick/Dev/Python-In-Action/placemap/chromedriver"
    chrome_driver = webdriver.Chrome(executable_path=driver_path,
                                     options=chrome_options)
    chrome_driver.set_page_load_timeout(120)
    chrome_driver.set_script_timeout(60)
    chrome_driver.implicitly_wait(30)
    return chrome_driver


def __init_crawl_job(countrys: List[str], matches: List[str]) -> None:
    chrome_driver = None
    try:
        print("start crawl data ......")
        chrome_driver = __init_chrome_dirver()
        chrome_driver.get(start_url)
        country_list: List[Country] = __parse_country_list(
            chrome_driver, countrys)
        for country in country_list:
            (country_name, country_href) = country
            print(f"国家[{country_name}]详情链接:{country_href}")
            chrome_driver.get(country_href)
            area_list: List[Area] = __parse_area_list(chrome_driver)
            for area in area_list:
                (area_name, area_href) = area
                print(f"国家[{country_name}] 地区[{area_name}]详情链接:{area_href}")
                chrome_driver.get(area_href)
                for pattern in matches:
                    target_href = str(pattern).lower().replace(" ", "-")
                    if pattern not in chrome_driver.page_source:
                        print(f"地区[{area_name}]不存在[{pattern}]信息")
                        continue
                    target_href = area_href + target_href
                    print(f"地区[{area_name}]的[{pattern}]详情链接:{target_href}")
                    chrome_driver.get(target_href)
                    match = re.search(page_pattern, chrome_driver.page_source)
                    if match == None:
                        continue
                    page_size = int(match.group(2))
                    # page_size = 3 if page_size > 3 else page_size
                    print(f"共有结果 [{page_size}] 页 ")
                    for page_num in range(1, page_size + 1):
                        p_el_list = __parse_taget_list(chrome_driver,
                                                       target_href, page_num)
                        for p_el in p_el_list:
                            try:
                                position = __parse_position(
                                    country_name, area_name, p_el)
                                process__queue.put(position, False, 30)
                                print("\n".join(position))
                            except:
                                print(sys.exc_info())
                                continue

    except:
        print("爬虫执行异常=", sys.exc_info())
    finally:
        global is_finishing
        is_finishing = True
        if chrome_driver is not None:
            chrome_driver.quit()


def __parse_country_list(chrome_driver: WebDriver,
                         countrys: List[str]) -> List[Country]:
    all_country_el_list: List[
        WebElement] = chrome_driver.find_elements_by_tag_name("a")
    country_el_list: List[WebElement] = list(
        filter(lambda a: a.text in countrys, all_country_el_list))
    if len(country_el_list) == 0:
        raise ValueError(f"没有找到对应国家链接信息")
    country_list = list(
        map(lambda el: (el.text, el.get_attribute("href")), country_el_list))
    return country_list


def __parse_area_list(chrome_driver: WebDriver) -> List[Area]:
    area_el_list = chrome_driver.find_elements_by_xpath(
        '//div[@class="four columns"]/a')
    return list(
        map(lambda el: (el.text, el.get_attribute("href")), area_el_list))


def __parse_position(country_name: str, area_name: str,
                     p_el: WebElement) -> Position:
    title = p_el.find_element_by_xpath("./b/a").text
    cordinate = p_el.find_element_by_xpath("./a").text
    cordinate_match = re.match(cordinate_pattern, cordinate)
    latitude = cordinate_match.group(1) if cordinate_match else ""
    longitude = cordinate_match.group(2) if cordinate_match else ""
    detail_list = str(p_el.text).split("\n")
    address = detail_list[1]
    detail = detail_list[3] if len(detail_list) == 4 else ""
    match_p = re.search(phone_pattern, detail)
    match_w = re.search(web_pattern, detail)
    phone = "" if not match_p else match_p.group(1).strip()
    web = "" if not match_w else match_w.group(1).strip()
    return (country_name, area_name, title, address, latitude, longitude,
            phone, web)


def __parse_taget_list(chrome_driver: WebDriver, target_href: str,
                       page_num: int) -> List[WebElement]:
    page_href = target_href + f"/{page_num}/"
    print(f"页面链接:{page_href}")
    chrome_driver.get(page_href)
    p_el_list: List[WebElement] = chrome_driver.find_elements_by_xpath(p_xpath)
    return list(filter(lambda el: str(el.text).strip() != '', p_el_list))


def __init_xls(country_list: List[str], columns: List[str]) -> Workbook:
    print("正在初始化表格...")
    workbook = None
    if os.path.exists(xls_file_path):
        workbook = load_workbook(xls_file_path)
    else:
        workbook = Workbook()
        workbook.remove(workbook.get_active_sheet())
        for country_name in country_list:
            worksheet = workbook.create_sheet(country_name)
            worksheet.sheet_properties.tabColor = '6FB7B7'
            for xls_col in worksheet.iter_cols(max_col=len(columns),
                                               max_row=1):
                for cell in xls_col:
                    cell.value = columns[cell.col_idx - 1]
    return workbook


def __write_xls(workbook: Workbook, position: Position) -> None:
    print("正在写入数据...")
    worksheet = workbook.get_sheet_by_name(position[0])
    worksheet.append(position[1:])


def __close_xls(workbook: Workbook, filename: str) -> None:
    workbook.save(filename)
    print("数据文件已保存.")


def __start_process_task(country_list: List[str],
                         columns: List[str]) -> threading.Thread:
    print("正在运行数据处理任务...")
    workbook = __init_xls(country_list, columns)
    thread = threading.Thread(target=__process_position,
                              name="position_task_0",
                              args=(workbook, ))
    return thread


def __process_position(workbook: Workbook) -> None:
    while not is_finishing:
        while not process__queue.empty():
            position = process__queue.get()
            __write_xls(workbook, position)
    __close_xls(workbook, xls_file_path)
